#!/usr/bin/env python3
"""
CONSULTA DE PROCESSOS EXECUTIVOS – COMPRA DE DÍVIDA
Tribunais: TJSP e TJMS
Credor: Banco do Brasil S/A
Valor mínimo: R$ 300.000

Dependências:
  pip install requests beautifulsoup4 openpyxl lxml
"""

import re
import json
import time
import unicodedata
from datetime import date
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════════════

CREDOR_PADRAO = "Banco do Brasil S.A."
VALOR_MINIMO  = 300_000.0
PAUSA         = 0.4
TIMEOUT       = 20

HEADERS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9",
}

COLUNAS_EXCEL = [
    "Número do Processo", "Tribunal", "Estado", "Foro / Comarca", "Vara",
    "Fase", "Situação Atual", "Leilão / Hasta Pública",
    "Credor", "Devedor", "Valor da Execução",
    "Tipo de Garantia", "Descrição da Garantia",
    "Tipo de Imóvel", "Localização do Imóvel", "Matrícula do Imóvel",
    "Prioridade", "Possível Compra",
    "Motivo da Prioridade", "Próximos Passos", "Observações"
]

CORES = {
    "alta":      "C6EFCE",
    "media":     "FFEB9C",
    "baixa":     "DDEEFF",
    "descartar": "FFC7CE",
}

CLASSES_ALVO = [
    "Execução de Título Extrajudicial",
    "Cumprimento de sentença",
    "Monitória",
    "Execução Fiscal",
]

# ═══════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ═══════════════════════════════════════════════════════════════════

def normalizar(texto):
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto.lower())
        if unicodedata.category(c) != 'Mn'
    )

def slug(texto):
    return normalizar(texto).replace(' ', '_')

def valor_para_float(valor_str):
    if not valor_str:
        return 0.0
    limpo = re.sub(r'[R$\s]', '', valor_str).replace('.', '').replace(',', '.')
    try:
        return float(limpo)
    except ValueError:
        return 0.0

# ═══════════════════════════════════════════════════════════════════
# BUSCAR CÓDIGO DO FORO AUTOMATICAMENTE
# ═══════════════════════════════════════════════════════════════════

def extrair_locais_pesquisa(html):
    """
    A lista de foros do e-SAJ não vem num <select> estático: é injetada
    via JS na própria página, num bloco como:
        $.saj.cpopg.locaisDePesquisa = JSON.parse('[{"id":318,"text":"Foro de Leme"}, ...]');
    Extrai e retorna essa lista (id, text) sem precisar executar JS.
    """
    m = re.search(r'locaisDePesquisa\s*=\s*JSON\.parse\(\'(.+?)\'\)\s*;', html, re.S)
    if not m:
        return []
    try:
        return json.loads(m.group(1))
    except (ValueError, json.JSONDecodeError):
        return []


def _melhor_foro(cidade, locais):
    """Escolhe, dentre a lista de foros, o que melhor casa com o nome da cidade."""
    cidade_norm = normalizar(cidade)
    melhor_opcao = None
    melhor_score = 0

    for item in locais:
        foro_id = item.get("id")
        nome    = item.get("text", "")
        nome_norm = normalizar(nome)

        if foro_id is None or foro_id == -1:
            continue

        if cidade_norm in nome_norm:
            score = len(cidade_norm)
            if nome_norm == cidade_norm or nome_norm == f"foro de {cidade_norm}":
                score += 100
            if score > melhor_score:
                melhor_score = score
                melhor_opcao = (str(foro_id), nome)

    return melhor_opcao if melhor_opcao else (None, None)


def buscar_foro_tjsp(cidade, session):
    """
    Busca o código do foro no e-SAJ TJSP pelo nome da cidade.
    Retorna (codigo_foro, nome_foro) ou (None, None).
    """
    url = "https://esaj.tjsp.jus.br/cpopg/open.do"
    try:
        html = session.get(url, timeout=TIMEOUT).text
    except Exception as e:
        print(f"  ⚠️  Erro ao carregar página do TJSP: {e}")
        return None, None

    locais = extrair_locais_pesquisa(html)
    if not locais:
        print("  ⚠️  Não encontrei a lista de foros no TJSP.")
        return None, None

    return _melhor_foro(cidade, locais)


def buscar_foro_tjms(cidade, session):
    """
    Busca o código do foro no e-SAJ TJMS pelo nome da cidade.
    Retorna (codigo_foro, nome_foro) ou (None, None).
    """
    url = "https://esaj.tjms.jus.br/cpopg5/open.do"
    try:
        html = session.get(url, timeout=TIMEOUT).text
    except Exception as e:
        print(f"  ⚠️  Erro ao carregar página do TJMS: {e}")
        return None, None

    locais = extrair_locais_pesquisa(html)
    if not locais:
        print("  ⚠️  Não encontrei a lista de foros no TJMS.")
        return None, None

    return _melhor_foro(cidade, locais)


def encontrar_tribunal_e_foro(cidade):
    """
    Procura a cidade no TJSP e no TJMS automaticamente.
    Retorna (tribunal, base_url, codigo_foro, nome_foro, estado) ou None.
    """
    session = requests.Session()
    session.headers.update(HEADERS_HTTP)

    print(f"  🔎 Procurando foro de '{cidade}' no TJSP...")
    foro, nome = buscar_foro_tjsp(cidade, session)
    if foro:
        print(f"  ✅ Encontrado no TJSP: {nome} (foro {foro})")
        return "TJSP", "https://esaj.tjsp.jus.br/cpopg", foro, nome, "SP", session

    print(f"  🔎 Não encontrado no TJSP. Procurando no TJMS...")
    foro, nome = buscar_foro_tjms(cidade, session)
    if foro:
        print(f"  ✅ Encontrado no TJMS: {nome} (foro {foro})")
        return "TJMS", "https://esaj.tjms.jus.br/cpopg5", foro, nome, "MS", session

    print(f"  ❌ Cidade '{cidade}' não encontrada em nenhum tribunal.")
    print("     Verifique se o nome está correto (ex: 'Ribeirão Preto', 'Três Lagoas')")
    return None

# ═══════════════════════════════════════════════════════════════════
# COLETA DE PROCESSOS
# ═══════════════════════════════════════════════════════════════════

PAPEIS = [
    "Reqte", "Reqdo", "Exqte", "Exqdo", "Credor",
    "Outros", "Interesdo.", "Interessado", "Embargante",
]

def parsear_pagina(html, base_url):
    soup = BeautifulSoup(html, "lxml")
    processos = []

    # Total
    total = 0
    for tag in soup.find_all(string=True):
        m = re.search(r'(\d+)\s+Processos?\s+encontrados?', str(tag), re.IGNORECASE)
        if m:
            total = int(m.group(1))
            break

    for link in soup.find_all("a", href=re.compile(r"show\.do")):
        numero = link.get_text(strip=True)
        if not re.match(r'\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}', numero):
            continue

        href_raw = link.get("href", "")
        if href_raw.startswith("/"):
            dominio = re.match(r'https?://[^/]+', base_url)
            href = dominio.group(0) + href_raw if dominio else base_url + href_raw
        elif href_raw.startswith("http"):
            href = href_raw
        else:
            href = base_url + "/" + href_raw

        bloco = link.find_parent("tr") or link.find_parent("li") or link.find_parent("div")
        linhas = []
        if bloco:
            linhas = [l.strip() for l in bloco.get_text("\n").split("\n") if l.strip()]

        classe = assunto = vara = data = papel = ""
        for i, l in enumerate(linhas):
            if not classe and any(l.startswith(c) for c in CLASSES_ALVO):
                classe = l
            elif classe and not assunto:
                if (not any(l.startswith(p) for p in PAPEIS)
                        and "Recebido" not in l
                        and len(l) > 3
                        and not re.match(r'\d{2}/\d{2}/\d{4}', l)):
                    assunto = l
            if not papel and any(l.rstrip(":") == p for p in PAPEIS):
                papel = l.rstrip(":")
            if "Recebido em:" in l and i + 1 < len(linhas):
                data = linhas[i + 1]
            if "Vara" in l and not vara:
                vara = l

        processos.append({
            "numero": numero, "href": href,
            "classe": classe, "assunto": assunto,
            "vara": vara, "data": data, "papel": papel,
        })

    return processos, total


def coletar_lista(session, base_url, foro, nome_parte):
    print(f"\n  📋 Buscando '{nome_parte}' no foro {foro}...")

    params_base = {
        "cbPesquisa":                  "NMPARTE",
        "dadosConsulta.valorConsulta": nome_parte,
        "chNmCompleto":                "true",
        "cdForo":                      foro,
    }

    try:
        html = session.get(
            f"{base_url}/search.do",
            params={**params_base, "paginaConsulta": 1},
            timeout=TIMEOUT
        ).text
    except Exception as e:
        print(f"  ❌ Erro na busca: {e}")
        return []

    todos, total = parsear_pagina(html, base_url)

    if total == 0:
        m = re.search(r'(\d+)\s+Processo', html)
        if m:
            total = int(m.group(1))

    total_pags = max(1, (total + 24) // 25)
    print(f"  → {total} processos / {total_pags} páginas")

    for pag in range(2, total_pags + 1):
        time.sleep(PAUSA)
        try:
            html = session.get(
                f"{base_url}/search.do",
                params={**params_base, "paginaConsulta": pag},
                timeout=TIMEOUT
            ).text
            procs, _ = parsear_pagina(html, base_url)
            todos.extend(procs)
        except Exception:
            pass
        print(f"  Pág {pag}/{total_pags} — {len(todos)} coletados", end="\r")

    print()
    return todos

# ═══════════════════════════════════════════════════════════════════
# DETALHAMENTO
# ═══════════════════════════════════════════════════════════════════

def detectar_garantia(texto):
    t = texto.lower()
    if re.search(r'penhor\s*(rural|de safra|de maquin|agr)', t):
        return "Penhor Rural (NÃO é imóvel)", "Penhor sobre maquinário/safra", False
    if re.search(r'c[eé]dula hipotec', t):
        return "Cédula Hipotecária", "Hipoteca via cédula hipotecária", True
    if re.search(r'hipotec', t):
        return "Hipoteca", "Garantia hipotecária sobre imóvel", True
    if re.search(r'aliena[cç][aã]o fiduci', t):
        return "Alienação Fiduciária", "Alienação fiduciária de imóvel", True
    if re.search(r'im[oó]vel penhorado|penhora de im[oó]vel', t):
        return "Penhora de Imóvel", "Imóvel penhorado nos autos", True
    if re.search(r'c[eé]dula de cr[eé]dito rural', t):
        return "Cédula de Crédito Rural", "CCR – verificar garantia nos autos", True
    if re.search(r'fazenda|gl[eê]ba|s[ií]tio|propriedade rural', t):
        return "Imóvel Rural (indício)", "Menção a imóvel rural", True
    if re.search(r'matr[ií]cula\s*n', t):
        return "Imóvel (matrícula)", "Matrícula de imóvel mencionada", True
    return "A verificar nos autos", "Sem indício público de garantia", False

def detectar_tipo_imovel(texto):
    t = texto.lower()
    if re.search(r'fazenda|gl[eê]ba|s[ií]tio|rural|ch[áa]cara', t):
        return "rural"
    if re.search(r'apartamento|urbano|sala comercial|loja|galp[aã]o|terreno urbano', t):
        return "urbano"
    return "A verificar"

def detalhar(session, href):
    try:
        r = session.get(href, timeout=TIMEOUT)
        html = r.text
    except Exception as e:
        return {"erro": str(e)}

    soup = BeautifulSoup(html, "lxml")
    texto = soup.get_text("\n", strip=True)

    # Situação
    situacao = ""
    for palavra in ["Ativo", "Suspenso", "Extinto", "Arquivado", "Baixado"]:
        if re.search(rf'\b{palavra}\b', texto, re.IGNORECASE):
            situacao = palavra
            break

    # Valor
    valor = ""
    bloco_mais = soup.find(id="maisDetalhes")
    texto_mais = bloco_mais.get_text("\n") if bloco_mais else texto
    m = re.search(r'Valor da a[çc][ãa]o\s*\n?\s*(R\$[\d\s.,]+)', texto_mais)
    if m:
        valor = m.group(1).strip()
    if not valor:
        ms = re.findall(r'R\$\s*[\d.,]+', texto_mais)
        if ms:
            valor = ms[0].strip()

    # Distribuição
    distribuicao = ""
    m = re.search(r'Distribui[çc][ãa]o\s*\n?\s*(\d{2}/\d{2}/\d{4})', texto_mais)
    if m:
        distribuicao = m.group(1)

    # Juiz
    juiz = ""
    m = re.search(r'\bJu[íi]z[a]?\b\s*\n?\s*([A-ZÁÀÂÃÉÊÍÓÔÕÚ][^\n]{3,60})', texto)
    if m:
        juiz = m.group(1).strip()

    # Partes
    exeqte = executado = ""
    tabela = (soup.find("table", id=re.compile(r"tableTodasPartes", re.I))
              or soup.find("div", id=re.compile(r"partes", re.I)))
    if tabela:
        linhas_p = [l.strip() for l in tabela.get_text("\n").split("\n") if l.strip()]
        for i, l in enumerate(linhas_p):
            if l in ("Exeqte", "Reqte", "Credor") and not exeqte and i+1 < len(linhas_p):
                exeqte = linhas_p[i+1]
            if l in ("Exectdo","Exectda","Reqdo","Reqda") and not executado and i+1 < len(linhas_p):
                executado = linhas_p[i+1]
    if not exeqte:
        m = re.search(r'(?:Exeqte|Reqte|Credor)\s+([A-ZÁÀÂÃÉÊÍÓÔÕÚ][^\n]+)', texto)
        if m:
            exeqte = m.group(1).strip().split('\n')[0]
    if not executado:
        m = re.search(r'(?:Exectdo|Exectda|Reqdo|Reqda)\s+([A-ZÁÀÂÃÉÊÍÓÔÕÚ][^\n]+)', texto)
        if m:
            executado = m.group(1).strip().split('\n')[0]

    # Movimentações
    ultima_mov = penultima_mov = ""
    tbody = soup.find("tbody", id=re.compile(r"tabelaTodasMovimentacoes|movimentacoes", re.I))
    if tbody:
        rows = tbody.find_all("tr")
        if rows:
            tds = rows[0].find_all("td")
            if len(tds) >= 2:
                ultima_mov = tds[0].get_text(strip=True) + " – " + tds[1].get_text(strip=True)[:100]
        if len(rows) > 1:
            tds2 = rows[1].find_all("td")
            if len(tds2) >= 2:
                penultima_mov = tds2[0].get_text(strip=True) + " – " + tds2[1].get_text(strip=True)[:100]
    if not ultima_mov:
        movs = re.findall(r'(\d{2}/\d{2}/\d{4})\s+([A-ZÁÀÂÃÉÊÍÓÔÕÚ][^\n]{5,80})', texto)
        if movs:
            ultima_mov = f"{movs[0][0]} – {movs[0][1]}"
        if len(movs) > 1:
            penultima_mov = f"{movs[1][0]} – {movs[1][1]}"

    # Leilão
    leilao = "Não"
    if re.search(r'hasta p[úu]blica|leil[ãa]o judicial|edital de leil', texto, re.IGNORECASE):
        m_data = re.search(r'(?:hasta|leil[ãa]o)[^.]{0,60}(\d{2}/\d{2}/\d{4})', texto, re.IGNORECASE)
        leilao = f"Sim – {m_data.group(1)}" if m_data else "Sim – data a confirmar"

    # Garantia
    tipo_g, desc_g, tem_imovel = detectar_garantia(texto)
    tipo_imovel = detectar_tipo_imovel(texto) if tem_imovel else "N/A"

    # Matrícula
    matricula = "N/A"
    m_mat = re.search(r'matr[íi]cula\s*(?:n[ºo°.]?\s*)?(\d[\d./\-]+)', texto, re.IGNORECASE)
    if m_mat:
        matricula = m_mat.group(1).strip()

    return {
        "situacao": situacao or "A verificar",
        "valor": valor,
        "valor_num": valor_para_float(valor),
        "distribuicao": distribuicao,
        "juiz": juiz,
        "exeqte": exeqte,
        "executado": executado,
        "ultima_mov": ultima_mov,
        "penultima_mov": penultima_mov,
        "leilao": leilao,
        "tipo_garantia": tipo_g,
        "desc_garantia": desc_g,
        "tem_imovel": tem_imovel,
        "tipo_imovel": tipo_imovel,
        "matricula": matricula,
    }

# ═══════════════════════════════════════════════════════════════════
# CLASSIFICAÇÃO
# ═══════════════════════════════════════════════════════════════════

def classificar(proc, det):
    classe     = proc.get("classe", "")
    valor      = det.get("valor_num", 0)
    situacao   = det.get("situacao", "").lower()
    leilao     = det.get("leilao", "Não")
    tipo_g     = det.get("tipo_garantia", "")
    tem_imovel = det.get("tem_imovel", False)
    exeqte     = det.get("exeqte", "")

    # Descartar
    if any(x in situacao for x in ["extinto", "arquivado", "baixado", "encerrado"]):
        return "descartar", False, f"❌ Processo {situacao}.", "Descartado."

    if 0 < valor < VALOR_MINIMO:
        return "descartar", False, f"❌ Valor R${valor:,.2f} abaixo de R${VALOR_MINIMO:,.0f}.", "Descartado."

    if "NÃO é imóvel" in tipo_g:
        return "descartar", False, "❌ Penhor de maquinário/safra, não imóvel.", "Descartado."

    bb_credor = bool(exeqte) and bool(re.search(r'banco do brasil', exeqte, re.I))
    if exeqte and not bb_credor:
        return "descartar", False, f"❌ Credor não é BB (é: {exeqte[:50]}).", "Descartado."

    # Alta
    if ("Execução de Título Extrajudicial" in classe
            and valor >= VALOR_MINIMO
            and tem_imovel
            and bb_credor):
        motivo = (f"✅ ALTA: Exec. Título Extrajudicial | BB credor | "
                  f"Valor R${valor:,.2f} | Garantia: {tipo_g}")
        passos = ("1) Acessar autos com login OAB – verificar seção 'DAS GARANTIAS' do CCB; "
                  "2) Confirmar matrícula e valor do imóvel; "
                  "3) Contatar advogado do BB para negociar cessão.")
        if "Sim" in leilao:
            passos = f"⚡ HASTA PÚBLICA: {leilao}. " + passos
        return "alta", True, motivo, passos

    # Média
    if "Execução de Título Extrajudicial" in classe and valor >= VALOR_MINIMO:
        motivo = (f"🟡 MÉDIA: Exec. Título Extrajudicial | Valor R${valor:,.2f} | "
                  f"Garantia não confirmada publicamente.")
        passos = ("1) Verificar garantia nos autos (login OAB); "
                  "2) Se hipoteca/AF confirmada → reclassificar como ALTA.")
        return "media", False, motivo, passos

    if "Execução de Título Extrajudicial" in classe and valor == 0:
        return "media", False, "🟡 MÉDIA: Valor não capturado – verificar manualmente.", \
               "Acessar processo no e-SAJ para confirmar valor e garantia."

    if "Cumprimento de sentença" in classe and valor >= VALOR_MINIMO and tem_imovel:
        motivo = f"🟡 MÉDIA: Cumprimento de sentença | Valor R${valor:,.2f} | Garantia: {tipo_g}"
        return "media", False, motivo, "Verificar penhora de imóvel e viabilidade de cessão."

    return "descartar", False, f"❌ Fora do escopo – classe '{classe}', valor R${valor:,.2f}.", "Descartado."

# ═══════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ═══════════════════════════════════════════════════════════════════

def varrer_cidade(cidade):
    print(f"\n{'='*60}")
    print(f"🔍 CIDADE: {cidade}")
    print(f"   Credor: {CREDOR_PADRAO}")
    print(f"   Valor mínimo: R$ {VALOR_MINIMO:,.0f}")
    print(f"{'='*60}")

    # Buscar foro automaticamente
    resultado = encontrar_tribunal_e_foro(cidade)
    if not resultado:
        return []

    tribunal, base_url, foro, nome_foro, estado, session = resultado

    # Coletar lista
    todos = coletar_lista(session, base_url, foro, CREDOR_PADRAO)
    print(f"\n  📦 {len(todos)} processos coletados")

    # Filtrar classes relevantes
    relevantes = [p for p in todos if any(c in p.get("classe","") for c in CLASSES_ALVO)]
    print(f"  🎯 {len(relevantes)} nas classes-alvo")

    if not relevantes:
        print("  ℹ️  Nenhum processo nas classes-alvo.")
        return []

    # Detalhar cada processo
    resultados = []
    for i, proc in enumerate(relevantes, 1):
        print(f"  [{i:>3}/{len(relevantes)}] {proc['numero']} | {proc.get('classe','')} | {proc.get('assunto','')}")
        time.sleep(PAUSA)

        det = detalhar(session, proc["href"])
        if "erro" in det:
            print(f"            ⚠️  {det['erro']}")
            det = {
                "situacao": "Erro", "valor": "", "valor_num": 0,
                "distribuicao": "", "juiz": "", "exeqte": "", "executado": "",
                "ultima_mov": "", "penultima_mov": "", "leilao": "Não",
                "tipo_garantia": "A verificar", "desc_garantia": "Erro ao acessar",
                "tem_imovel": False, "tipo_imovel": "A verificar", "matricula": "N/A",
            }

        pri, compra, motivo, passos = classificar(proc, det)
        emoji = {"alta":"🟢","media":"🟡","baixa":"🔵","descartar":"🔴"}.get(pri,"⚪")
        print(f"            {emoji} {pri.upper()} | {det.get('valor','sem valor')} | {det.get('situacao','')}")

        resultados.append({
            "numero_processo":    proc["numero"],
            "tribunal":           tribunal,
            "estado":             estado,
            "foro_comarca":       nome_foro,
            "vara":               proc.get("vara",""),
            "fase":               proc.get("classe","") + (" – " + proc.get("assunto","") if proc.get("assunto") else ""),
            "situacao_atual":     det.get("ultima_mov","") or det.get("situacao",""),
            "leilao":             det.get("leilao","Não"),
            "credor":             det.get("exeqte","") or CREDOR_PADRAO,
            "devedor":            det.get("executado",""),
            "valor_execucao":     det.get("valor",""),
            "valor_num":          det.get("valor_num",0),
            "tipo_garantia":      det.get("tipo_garantia","A verificar"),
            "descricao_garantia": det.get("desc_garantia",""),
            "tipo_imovel":        det.get("tipo_imovel","A verificar"),
            "localizacao_imovel": f"{cidade} / {estado}",
            "matricula_imovel":   det.get("matricula","N/A"),
            "prioridade":         pri,
            "possivel_compra":    compra,
            "motivo_prioridade":  motivo,
            "proximos_passos":    passos,
            "observacoes": (
                f"Juiz: {det.get('juiz','')}. "
                f"Distribuído: {det.get('distribuicao','')}. "
                f"Penúlt. mov.: {det.get('penultima_mov','')}."
            ).strip(),
        })

    # Ordenar
    ordem = {"alta":0,"media":1,"baixa":2,"descartar":3}
    resultados.sort(key=lambda r: (ordem.get(r["prioridade"],9), -r.get("valor_num",0)))

    alta  = sum(1 for r in resultados if r["prioridade"]=="alta")
    media = sum(1 for r in resultados if r["prioridade"]=="media")
    print(f"\n  📊 {len(resultados)} analisados | 🟢 Alta: {alta} | 🟡 Média: {media}")

    return resultados

# ═══════════════════════════════════════════════════════════════════
# GERAR EXCEL
# ═══════════════════════════════════════════════════════════════════

def gerar_excel(cidade, resultados, pasta="."):
    if not resultados:
        return ""

    hoje     = date.today().strftime("%d%m%Y")
    caminho  = Path(pasta) / f"processos_{slug(cidade)}_{hoje}.xlsx"
    wb       = openpyxl.Workbook()

    # ── Aba 1: Processos ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Processos"

    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 38

    for col, h in enumerate(COLUNAS_EXCEL, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align

    widths = [32,8,6,24,22,32,55,28,30,30,18,28,55,14,22,18,12,14,55,60,55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rfont = Font(name="Arial", size=9)
    wrap  = Alignment(wrap_text=True, vertical="top")

    for ri, proc in enumerate(resultados, 2):
        cor = CORES.get(proc["prioridade"], "FFFFFF")
        ws.row_dimensions[ri].height = 70
        vals = [
            proc.get("numero_processo",""), proc.get("tribunal",""),
            proc.get("estado",""), proc.get("foro_comarca",""),
            proc.get("vara",""), proc.get("fase",""),
            proc.get("situacao_atual",""), proc.get("leilao","Não"),
            proc.get("credor",""), proc.get("devedor",""),
            proc.get("valor_execucao",""), proc.get("tipo_garantia",""),
            proc.get("descricao_garantia",""), proc.get("tipo_imovel",""),
            proc.get("localizacao_imovel",""), proc.get("matricula_imovel",""),
            proc.get("prioridade","").upper(),
            "Sim" if proc.get("possivel_compra") else "Não",
            proc.get("motivo_prioridade",""), proc.get("proximos_passos",""),
            proc.get("observacoes",""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = PatternFill("solid", fgColor=cor)
            cell.font = rfont; cell.alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS_EXCEL))}{len(resultados)+1}"

    # ── Aba 2: Resumo ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo Executivo")
    ws2.column_dimensions["A"].width = 44
    ws2.column_dimensions["B"].width = 35

    def sc(row, col, val, bold=False, size=10, bg=None, merge_to=None):
        c = ws2.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, name="Arial", size=size)
        c.alignment = Alignment(horizontal="center" if merge_to else "left",
                                vertical="center", wrap_text=True)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        if merge_to: ws2.merge_cells(f"A{row}:{get_column_letter(merge_to)}{row}")

    tribunal_nome = resultados[0].get("tribunal","")
    sc(1,1,f"VARREDURA – {cidade.upper()} / {tribunal_nome}",
       bold=True, size=13, bg="D9E1F2", merge_to=2)
    ws2.row_dimensions[1].height = 30

    alta  = [r for r in resultados if r["prioridade"]=="alta"]
    media = [r for r in resultados if r["prioridade"]=="media"]

    for i, (k, v) in enumerate([
        ("Data",               date.today().strftime("%d/%m/%Y")),
        ("Tribunal",           tribunal_nome),
        ("Cidade",             cidade),
        ("Credor",             CREDOR_PADRAO),
        ("Valor mínimo",       f"R$ {VALOR_MINIMO:,.0f}"),
        ("Total analisados",   str(len(resultados))),
        ("🟢 Alta",            str(len(alta))),
        ("🟡 Média",           str(len(media))),
        ("✅ Possível compra", str(sum(1 for r in resultados if r.get("possivel_compra")))),
        ("⚡ Com leilão",      str(sum(1 for r in resultados if "Sim" in r.get("leilao","")))),
    ], 2):
        sc(i,1,k,bold=True); sc(i,2,v)
        ws2.row_dimensions[i].height = 18

    r = 13
    sc(r,1,"⭐ TOP OPORTUNIDADES", bold=True, size=11, bg="C6EFCE", merge_to=2)
    ws2.row_dimensions[r].height = 22; r += 1

    for idx, proc in enumerate((alta+media)[:5], 1):
        txt = (f"{idx}. {proc['numero_processo']} | "
               f"{proc.get('valor_execucao','')} | "
               f"{proc.get('devedor','')} | {proc.get('prioridade','').upper()}")
        sc(r,1,txt, merge_to=2)
        ws2.cell(row=r,column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r].height = 28; r += 1

    r += 1
    sc(r,1,"⚠️ ALERTAS", bold=True, size=11, bg="FFEB9C", merge_to=2)
    ws2.row_dimensions[r].height = 22; r += 1
    for alerta in [
        "• Valores do portal público podem diferir do valor atualizado nos autos.",
        "• Garantias identificadas por palavras-chave – confirmar com login OAB.",
        "• BB raramente cede crédito individualmente – negociar via advogado.",
        "• Processos com valor R$0 tiveram falha na captura – verificar manualmente.",
    ]:
        sc(r,1,alerta, merge_to=2)
        ws2.cell(row=r,column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r].height = 22; r += 1

    # ── Aba 3: Foco ──────────────────────────────────────────────────
    foco = alta + media
    if foco:
        ws3 = wb.create_sheet("⭐ Foco")
        h3 = ["Nº Processo","Devedor","Valor","Garantia","Fase","Prioridade","Próximos Passos"]
        w3 = [32, 30, 18, 30, 32, 12, 65]
        for ci, (h, w) in enumerate(zip(h3, w3), 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor="375623")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.row_dimensions[1].height = 28
        for ri, proc in enumerate(foco, 2):
            cor = CORES.get(proc["prioridade"],"FFFFFF")
            ws3.row_dimensions[ri].height = 60
            for ci, v in enumerate([
                proc.get("numero_processo",""), proc.get("devedor",""),
                proc.get("valor_execucao",""), proc.get("tipo_garantia",""),
                proc.get("fase",""), proc.get("prioridade","").upper(),
                proc.get("proximos_passos",""),
            ], 1):
                cell = ws3.cell(row=ri, column=ci, value=v)
                cell.fill = PatternFill("solid", fgColor=cor)
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(caminho)
    print(f"\n  💾 Excel salvo: {caminho}")
    return str(caminho)

# ═══════════════════════════════════════════════════════════════════
# MODO INTERATIVO
# ═══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  CONSULTA DE PROCESSOS EXECUTIVOS – COMPRA DE DÍVIDA")
    print(f"  Credor: {CREDOR_PADRAO}")
    print(f"  Valor mínimo: R$ {VALOR_MINIMO:,.0f}")
    print("=" * 60)
    print()
    print("  Digite o nome de qualquer cidade do Brasil.")
    print("  O sistema buscará automaticamente no TJSP ou TJMS.")
    print()
    print("  Comandos especiais:")
    print("    sair → encerrar o programa")
    print()

    pasta_saida = str(Path(__file__).parent)

    while True:
        cidade = input("  Cidade: ").strip()

        if not cidade:
            continue

        if cidade.lower() == "sair":
            print("\n  Até logo!")
            break

        try:
            resultados = varrer_cidade(cidade)
            if resultados:
                gerar_excel(cidade, resultados, pasta_saida)
            else:
                print(f"\n  ℹ️  Nenhum processo encontrado para '{cidade}'.")
        except KeyboardInterrupt:
            print("\n\n  Interrompido. Digite 'sair' para encerrar.")
        except Exception as e:
            print(f"\n  ❌ Erro: {e}")
            import traceback; traceback.print_exc()

        print("\n" + "-" * 60)


if __name__ == "__main__":
    main()